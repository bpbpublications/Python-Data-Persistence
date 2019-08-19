#example 10.8
from openpyxl.formula.translate import Translator#copy formula
sheet1['D2']='DIFF'
sheet1['D3']='=C$9-C3'
sheet1['D4'] = Translator("=C$9-C3", origin="D3").translate_formula("D4")
for row in range(4,8):
        coor=sheet1.cell(column=4, row=row).coordinate#copy formula to range
        sheet1.cell(column=4, row=row).value=Translator("=C$9-C3", origin="D3"). \
                                              translate_formula(coor)

#example 10.9

sheet1['A10'] = '=A2'
for col in range(1,4):
        coor=sheet1.cell(column=col,row=3).coordinate
        coor1=sheet1.cell(column=col, row=10).coordinate
        print (coor,coor1)
        sheet1.cell(column=col, row=10).value=Translator("=A2", origin="A10"). \
                                              translate_formula(coor1)
