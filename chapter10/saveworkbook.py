#saveworkbook.py
#example 10.4
from openpyxl import Workbook

wb = Workbook()
sheet1 = wb.active
sheet1.title='PriceList'
sheet1.cell(column=1, row=1, value='Pricelist')
pricelist=[('ProductID', 'Name', 'Price'),
           (1,'Laptop',25000),(2, 'TV',40000),
           (3,'Router',2000),(4,'Scanner',5000),
           (5,'Printer',9000), (6,'Mobile',15000)]
for col in range(1,4):
        for row in range(1,7):
                sheet1.cell(column=col, row=1+row, value=pricelist[row-1][col-1])

wb.save(filename = "test.xlsx")
