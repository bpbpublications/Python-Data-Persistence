#example 10.7
#readworkbook.py
from openpyxl import load_workbook

wb = load_workbook('test.xlsx')
sheet1 = wb['PriceList']
#copy range to list
pricelist=[]
for row in range(1,7):
        prod=[]
        for col in range(1,4):
                val=sheet1.cell(column=col, row=2+row).value
                prod.append(val)
        pricelist.append(tuple(prod))
print (pricelist)
#merge and center
sheet1.merge_cells('A1:C1')#merge
cell=sheet1['A1']
cell.alignment=Alignment(horizontal='center')
#apply font
cell.font=Font(name='Calibri',size=20,bold=True)
#define formula
sheet1['b8']='SUM'
sheet1['C8']='=SUM(C3:C7)'
sheet1['C9']='=AVERAGE(C3:C7)'
